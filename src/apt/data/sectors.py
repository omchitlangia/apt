"""Parse sector/industry mapping from Merge_21May2021.xlsx.

Outputs a parquet with columns: symbol, industry, isin, company_name.
Optionally enriches with BSE sub-industry via ISIN match.
"""

from __future__ import annotations

from pathlib import Path

import polars as pl
from loguru import logger


def _normalise_symbol(s: str | None) -> str | None:
    """Uppercase and strip whitespace from a symbol string."""
    if s is None:
        return None
    return str(s).strip().upper()


def parse_nifty500_sectors(xlsx_path: Path) -> pl.DataFrame:
    """Parse ind_nifty500list sheet → DataFrame(symbol, company_name, industry, isin).

    - Symbols are uppercased and whitespace-stripped.
    - Rows with null symbol are dropped.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["ind_nifty500list"]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("ind_nifty500list sheet is empty")

    header = rows[0]
    # Expected columns: Company Name, Industry, Symbol, Series, ISIN Code, ...
    col_idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
    logger.debug("Nifty500 sheet columns: {}", list(col_idx.keys()))

    records = []
    for row in rows[1:]:
        symbol = _normalise_symbol(row[col_idx.get("Symbol", 2)])
        company = row[col_idx.get("Company Name", 0)]
        industry = row[col_idx.get("Industry", 1)]
        isin = row[col_idx.get("ISIN Code", 4)]

        if not symbol:
            continue

        records.append(
            {
                "symbol": symbol,
                "company_name": str(company).strip() if company else None,
                "industry": str(industry).strip().upper() if industry else None,
                "isin": str(isin).strip() if isin else None,
            }
        )

    df = pl.DataFrame(records)
    logger.info("Parsed {} symbols from ind_nifty500list", len(df))
    return df


def _parse_bse_sectors(xlsx_path: Path) -> pl.DataFrame:
    """Parse Equity_BSE_ListOfSecurities sheet for sub-industry enrichment via ISIN."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Equity_BSE_ListOfSecurities_21-"]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pl.DataFrame({"isin": [], "bse_industry": []})

    header = rows[0]
    col_idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

    # Columns: ISIN No (col 0 or 8), Industry (col 9), Security Id (col 3)
    isin_col = col_idx.get("ISIN No", 0)
    industry_col = col_idx.get("Industry", 9)

    records = []
    for row in rows[1:]:
        isin = row[isin_col] if isin_col < len(row) else None
        bse_ind = row[industry_col] if industry_col < len(row) else None
        if isin:
            records.append(
                {
                    "isin": str(isin).strip(),
                    "bse_industry": str(bse_ind).strip() if bse_ind else None,
                }
            )

    return pl.DataFrame(records).unique(subset=["isin"])


def build_sector_mapping(
    xlsx_path: Path,
    output_path: Path,
    universe_symbols: set[str] | None = None,
) -> pl.DataFrame:
    """Parse xlsx, optionally enrich with BSE sub-industry, write parquet.

    Args:
        xlsx_path: path to Merge_21May2021.xlsx
        output_path: destination parquet path
        universe_symbols: if provided, report which universe symbols lack a mapping.

    Returns the sector DataFrame.
    """
    nifty = parse_nifty500_sectors(xlsx_path)

    # Enrich with BSE sub-industry
    try:
        bse = _parse_bse_sectors(xlsx_path)
        nifty = nifty.join(bse, on="isin", how="left")
        logger.info(
            "Enriched {} symbols with BSE sub-industry", nifty["bse_industry"].drop_nulls().len()
        )
    except Exception as exc:
        logger.warning("BSE enrichment skipped: {}", exc)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    nifty.write_parquet(output_path, use_pyarrow=True)
    logger.info("Wrote sector mapping → {}", output_path)

    # Report unmapped symbols
    if universe_symbols:
        mapped = set(nifty["symbol"].to_list())
        missing = sorted(universe_symbols - mapped)
        if missing:
            logger.warning(
                "{} universe symbols without sector mapping: {}",
                len(missing),
                missing,
            )
        else:
            logger.info("All universe symbols have sector mapping")

    return nifty
